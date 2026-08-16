"""
ادرس في مصر - Agent
Streamlit App
"""

import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime
import requests
import openpyxl
import json
import io
import time
import random

# ==================== يوزرات المكاتب ====================
OFFICES = {
    "office1": "pass1234",
    "office2": "pass5678",
}

SHEET_ID = "1BlFdtY-7ZIF1y2GwVosxlG9r7nK5xqYeW6yiIjPI_9U"
DRIVE_FOLDER_ID = "12L_qSHBnW4-tfQZRteynInWNBAML016f"

# ==================== Google Sheets Logging ====================

def get_sheet():
    """اتصل بـ Google Sheets"""
    try:
        creds_dict = st.secrets["gcp_service_account"]
        creds = Credentials.from_service_account_info(
            creds_dict,
            scopes=["https://spreadsheets.google.com/feeds",
                    "https://www.googleapis.com/auth/drive"]
        )
        client = gspread.authorize(creds)
        sheet = client.open_by_key(SHEET_ID).sheet1
        return sheet
    except Exception as e:
        print(f"خطأ في Google Sheets: {e}")
        return None

def log_to_sheet(office, action, filename=""):
    """بيسجل كل عملية في Google Sheets"""
    try:
        sheet = get_sheet()
        if sheet:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sheet.append_row([now, office, action, filename])
    except Exception as e:
        print(f"خطأ في التسجيل: {e}")

def upload_to_drive(file_bytes, filename, office):
    """بيرفع الإكسيل على Google Drive"""
    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseUpload
        import io as _io

        creds_dict = st.secrets["gcp_service_account"]
        creds = Credentials.from_service_account_info(
            creds_dict,
            scopes=["https://www.googleapis.com/auth/drive"]
        )

        service = build("drive", "v3", credentials=creds)

        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        drive_filename = f"{office}_{now}_{filename}"

        file_metadata = {
            "name": drive_filename,
            "parents": [DRIVE_FOLDER_ID]
        }

        media = MediaIoBaseUpload(
            _io.BytesIO(file_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id"
        ).execute()

        return True
    except Exception as e:
        print(f"خطأ في رفع الملف: {e}")
        return False

# ==================== API ====================
BASE_URL = "https://apiadm.study-in-egypt.gov.eg/api"
SITE_URL = "https://admission.study-in-egypt.gov.eg"

HEADERS_BASE = {
    "accept": "application/json, text/plain, */*",
    "accept-language": "ar",
    "device": "CITIZEN",
    "origin": SITE_URL,
    "referer": SITE_URL + "/",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "content-type": "application/json",
}


def human_delay(min_sec=2, max_sec=5):
    """استنى وقت عشوائي زي إنسان"""
    time.sleep(random.uniform(min_sec, max_sec))


def api_login(email, password):
    """بيعمل لوجين ويرجع session جديدة فيها الكوكي"""
    try:
        session = requests.Session()
        session.headers.update(HEADERS_BASE)

        # استنى شوية قبل اللوجين
        human_delay(2, 4)

        res = session.post(
            f"{BASE_URL}/student/login",
            json={"email": email, "password": password},
            timeout=30
        )

        if res.status_code not in [200, 201]:
            return None, None, f"فشل اللوجين - كود: {res.status_code}"

        # استنى شوية بعد اللوجين
        human_delay(2, 3)

        csrf_token = res.json().get("token", "") or res.headers.get("x-csrf-token", "")

        return session, csrf_token, None

    except Exception as e:
        return None, None, str(e)


def api_logout(session):
    """تسجيل خروج"""
    try:
        session.post(
            f"{BASE_URL}/student/logout",
            json={"redirectUrl": SITE_URL},
            timeout=15
        )
        human_delay(1, 2)
    except:
        pass


def get_status(session, csrf_token):
    """بيجيب حالة أحدث طلب"""
    try:
        filter_obj = {
            "where": {},
            "limit": 10,
            "offset": 0,
            "order": "statusUpdatedAt DESC",
            "fields": ["serviceSlug", "ID", "createdAt", "statusUpdatedAt", "activityId", "activityName"]
        }

        headers = {}
        if csrf_token:
            headers["x-csrf-token"] = csrf_token

        # استنى شوية قبل الـ request
        human_delay(1, 3)

        res = session.get(
            f"{BASE_URL}/dynamic_services/inbox",
            params={"filter": json.dumps(filter_obj)},
            headers=headers,
            timeout=30
        )

        if res.status_code not in [200, 304]:
            return "", f"خطأ ({res.status_code})"

        data = res.json()
        results = data.get("result", [])

        if not results:
            return "", "مفيش طلبات"

        # أحدث طلب (مرتبين من الأحدث للأقدم)
        latest = results[0]

        # جدول ترجمة activityName للعربي الصح
        translations = {
            "قبول الفحص الفنى": "القبول المبدئي",
            "قبول الفحص الفني": "القبول المبدئي",
            "kb8ijfo8": "تم السداد",
            "تم السداد": "تم السداد",
            "تأكيد استلام الملف وصحة و اكتمال المستندات": "تأكيد استلام الملف وصحة واكتمال المستندات",
            "الانتظار مراجعة الطلب": "بانتظار مراجعة الطلب",
            "قبول من رئيس الادارة المركزية": "قبول من رئيس الإدارة المركزية",
        }
        raw_status = latest.get("activityName", "غير محدد")
        status = translations.get(raw_status, raw_status)
        app_id = str(latest.get("ID", ""))
        return app_id, status

    except Exception as e:
        return "", f"خطأ: {e}"


def find_excel_columns(ws):
    cols = {"name": None, "email": None, "password": None, "status": None}
    header_row_num = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        row_values = [str(c).strip() if c else "" for c in row]
        if any("يميل" in v or "mail" in v.lower() for v in row_values):
            header_row_num = row_idx
            for i, cell in enumerate(row_values):
                cell_lower = cell.lower()
                if any(k in cell for k in ["اسم", "الإسم", "الاسم"]) or "name" in cell_lower:
                    cols["name"] = i
                elif any(k in cell for k in ["يميل", "بريد"]) or "mail" in cell_lower:
                    cols["email"] = i
                elif any(k in cell for k in ["باسورد", "كلمة المرور", "password", "pass"]) or "pass" in cell_lower:
                    cols["password"] = i
                elif any(k in cell for k in ["حالة", "الحالة", "status"]):
                    cols["status"] = i
            break
    if header_row_num is None:
        raise Exception("مش لاقي هيدر الإكسيل!")
    if cols["email"] is None:
        raise Exception("مش لاقي عمود الإيميل!")
    if cols["password"] is None:
        raise Exception("مش لاقي عمود الباسورد!")
    return cols, header_row_num


# ==================== الواجهة ====================

st.set_page_config(
    menu_items={"Get help": None, "Report a bug": None, "About": None},
    page_title="ادرس في مصر - Agent",
    page_icon="🎓",
    layout="centered"
)

st.markdown("""
<style>
    body { direction: rtl; }
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    [data-testid="stToolbar"] {visibility: hidden;}
    [data-testid="stDecoration"] {visibility: hidden;}
    [data-testid="stStatusWidget"] {visibility: hidden;}
    .viewerBadge_container__r5tak {display: none;}
    .styles_viewerBadge__CvC9N {display: none;}
    #stDecoration {display: none;}
    .stApp > header {display: none;}
    [data-testid="stSidebarNav"] {display: none;}
    .css-1jc7ptx, .e1ewe7hr3, .viewerBadge_container__1QSob,
    .styles_viewerBadge__1yB5_, .viewerBadge_link__1S137,
    .viewerBadge_text__1JaDK, div[class*="viewerBadge"],
    div[class*="StatusWidget"], [data-testid="stBottom"] > div:last-child {display: none !important;}
    .stApp { background-color: #0f1923; }
    h1 { color: #00c6ff !important; text-align: center; }
    .stButton > button {
        background-color: #00c6ff;
        color: #0f1923;
        font-weight: bold;
        border: none;
        border-radius: 8px;
        width: 100%;
    }
    .stButton > button:hover { background-color: #00a8d6; }
</style>
""", unsafe_allow_html=True)

st.markdown("<h1>🎓 ادرس في مصر</h1>", unsafe_allow_html=True)
st.markdown("<p style='text-align:center;color:#7a9cc0'>Agent تتبع حالة الطلبات</p>",
            unsafe_allow_html=True)
st.divider()

# ==================== لوجين ====================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    st.subheader("تسجيل الدخول")
    username = st.text_input("اسم المكتب")
    password = st.text_input("كلمة المرور", type="password")
    if st.button("دخول"):
        if OFFICES.get(username) == password:
            st.session_state.logged_in = True
            st.session_state.office = username
            log_to_sheet(username, "تسجيل دخول")
            st.rerun()
        else:
            st.error("اسم المكتب أو الباسورد غلط!")
    st.stop()

# ==================== الصفحة الرئيسية ====================
st.markdown(f"مرحباً بمكتب: **{st.session_state.office}**")

uploaded = st.file_uploader("ارفع ملف الإكسيل", type=["xlsx", "xls"])

if uploaded and st.button("▶ ابدأ"):
    log_to_sheet(st.session_state.office, "رفع ملف", uploaded.name)
    file_bytes = uploaded.read()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    try:
        cols, header_row_num = find_excel_columns(ws)
    except Exception as e:
        st.error(str(e))
        st.stop()

    log_area = st.empty()
    progress = st.progress(0)

    rows_data  = list(ws.iter_rows(min_row=header_row_num + 1, values_only=False))
    valid_rows = [r for r in rows_data
                  if r[cols["email"]].value and r[cols["password"]].value]
    total = len(valid_rows)

    log_lines = []
    success = failed = 0

    for idx, row in enumerate(valid_rows):
        email    = str(row[cols["email"]].value).strip()
        password = str(row[cols["password"]].value).strip()
        name     = row[cols["name"]].value if cols["name"] is not None else ""

        log_lines.append(f"👤 {name} | {email}")
        log_area.code("\n".join(log_lines))

        # لوجين
        session, csrf_token, err = api_login(email, password)

        if err or not session:
            log_lines.append(f"   ❌ فشل اللوجين: {err}\n")
            if cols["status"] is not None:
                row[cols["status"]].value = "فشل تسجيل الدخول"
            failed += 1
        else:
            # جيب الحالة
            app_num, status = get_status(session, csrf_token)
            if cols["status"] is not None:
                row[cols["status"]].value = status
            log_lines.append(f"   ✅ {status}")

            # تسجيل خروج
            api_logout(session)
            log_lines.append(f"   🚪 تم تسجيل الخروج\n")
            success += 1

        log_area.code("\n".join(log_lines))
        progress.progress((idx + 1) / total)

        # delay بين كل طالب والتاني (5-10 ثواني)
        if idx < total - 1:
            human_delay(5, 10)

    # احفظ الإكسيل
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    log_lines.append("─" * 40)
    log_lines.append(f"✅ خلصنا! إجمالي: {total} | نجح: {success} | فشل: {failed}")
    log_area.code("\n".join(log_lines))

    # ارفع الإكسيل على Drive
    upload_to_drive(out.getvalue(), uploaded.name, st.session_state.office)
    log_to_sheet(st.session_state.office, "اكتمل المعالجة", uploaded.name)

    st.success("خلصنا! حملي الإكسيل المحدث 👇")
    st.download_button(
        label="⬇ تحميل الإكسيل المحدث",
        data=out,
        file_name="students_updated.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if st.button("خروج"):
    log_to_sheet(st.session_state.get("office",""), "تسجيل خروج")
    st.session_state.clear()
    st.rerun()
